"""Build a review-friendly Excel workbook for the V4 classification draft."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image as PILImage
from PIL import ImageDraw, ImageFont

#: `docs/risk_scoring_v2.md` §3, in weight order. A reviewer scores each one
#: 0-4; the workbook never asks for a total, because the backend computes it and
#: a spreadsheet formula would be a second implementation of the rubric.
CRITERIA = (
    ("human_safety", "An toàn con người"),
    ("property_spread", "Mức lan thiệt hại"),
    ("essential_function", "Chức năng thiết yếu"),
    ("affected_scope", "Phạm vi căn hộ"),
    ("deterioration_speed", "Tốc độ xấu đi"),
)

#: The eleven named emergency facts. A reviewer picks from this list rather than
#: typing, because a blocker that does not match a code cannot floor anything.
BLOCKER_CODES = (
    "FIRE_OR_SMOKE",
    "ELECTRIC_SHOCK_OR_LIVE_WIRE",
    "GAS_LEAK_OR_ASPHYXIATION",
    "SERIOUS_INJURY",
    "PERSON_TRAPPED_IN_ELEVATOR",
    "SOLE_ESCAPE_ROUTE_BLOCKED",
    "ONGOING_VIOLENCE",
    "SEWAGE_OVERFLOW",
    "HEAVY_WATER_FLOW_SPREAD_RISK",
    "TOTAL_UNPLANNED_UTILITY_LOSS",
    "SOLE_TOILET_UNUSABLE",
)


def criterion_value(output, key):
    """What the draft says for one criterion, or blank.

    Blank on a v1-shaped draft, and deliberately not derived from `severity`:
    turning MEDIUM into five numbers would be inventing the judgements this
    workbook exists to collect.
    """
    criteria = output.get("criteria") or {}
    value = criteria.get(key)
    return "" if value is None else value


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE = REPO_ROOT / "Self_Dev_Docs" / "pairwise_v4" / "classification_golden_draft_v4.json"
DEFAULT_OUTPUT = REPO_ROOT / "outputs" / "golden_dataset_review" / "classification_golden_review_v4.xlsx"
DEFAULT_PREVIEW = REPO_ROOT / "outputs" / "golden_dataset_review" / "classification_golden_review_preview.png"

NAVY = "17324D"
BLUE = "246BCE"
LIGHT_BLUE = "EAF2FC"
TEAL = "0F766E"
LIGHT_TEAL = "DDF4EF"
AMBER = "D97706"
LIGHT_AMBER = "FFF4D6"
RED = "B42318"
LIGHT_RED = "FDE8E7"
GREEN = "16803A"
LIGHT_GREEN = "E3F6E8"
GRAY = "667085"
LIGHT_GRAY = "F2F4F7"
WHITE = "FFFFFF"
GRID = "D0D5DD"

THIN_GRAY = Side(style="thin", color=GRID)


def style_title(ws, cell_range: str, text: str, subtitle: str | None = None) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[cell.row].height = 34
    if subtitle:
        row = cell.row + 1
        end_col = ws[cell_range.split(":")[1]].column
        ws.merge_cells(start_row=row, start_column=cell.column, end_row=row, end_column=end_col)
        sub = ws.cell(row, cell.column, subtitle)
        sub.fill = PatternFill("solid", fgColor=NAVY)
        sub.font = Font(name="Aptos", size=10, italic=True, color="D7E5F3")
        sub.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22


def style_header(row) -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=BLUE))


def add_table(ws, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def fit_image(path: Path, max_width: int = 112, max_height: int = 82) -> ExcelImage:
    with PILImage.open(path) as source:
        width, height = source.size
    ratio = min(max_width / width, max_height / height)
    image = ExcelImage(path)
    image.width = max(1, int(width * ratio))
    image.height = max(1, int(height * ratio))
    return image


def flat_detail(case: dict[str, Any]) -> list[Any]:
    source = case["source"]
    dimensions = case["dimensions"]
    inputs = case["input"]
    output = case["expected_output"]
    return [
        case["case_id"],
        case["title"],
        case["review_status"],
        source["kind"],
        source.get("fixture_id"),
        dimensions["input_mode"],
        dimensions["image_state"],
        dimensions["evidence_relation"],
        dimensions["red_flag_source"],
        inputs["description"],
        "\n".join(inputs["image_paths"]),
        inputs["location_label"],
        inputs["floor_label"],
        inputs.get("unit_code"),
        inputs.get("confirmed_category"),
        json.dumps(inputs["conversation"], ensure_ascii=False),
        output["category"],
        output["text_category"],
        output["image_category"],
        *(criterion_value(output, key) for key, _label in CRITERIA),
        ",".join(output.get("blockers") or []),
        output["understandable"],
        output["image_relevant"],
        output["location_consistent"],
        "\n".join(output["incident_facts"]),
        output["ai_reason"],
        output["question_kind"],
        output["question_text"],
        ", ".join(output["category_options"] or []),
        "\n".join(case["review_notes"]),
    ]


def build_overview(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Tổng quan"
    ws.sheet_view.showGridLines = False
    style_title(
        ws,
        "A1:K1",
        "Golden Dataset V4 — Bảng review",
        "Bản nháp PENDING_HUMAN_REVIEW · Không dùng làm sealed test trước khi duyệt xong",
    )

    cards = [
        ("A4:B4", "A5:B6", "Tổng case", "=COUNTA('Review'!$A$2:$A$43)", BLUE),
        ("D4:E4", "D5:E6", "Case có ảnh", '=COUNTIF(\'Review\'!$F$2:$F$43,"Có ảnh")', TEAL),
        ("G4:H4", "G5:H6", "Red flag", '=COUNTIF(\'Review\'!$P$2:$P$43,TRUE)', RED),
        ("J4:K4", "J5:K6", "Cần xác nhận", '=COUNTIF(\'Review\'!$G$2:$G$43,"CÓ")', AMBER),
    ]
    for label_range, value_range, label, formula, color in cards:
        ws.merge_cells(label_range)
        ws.merge_cells(value_range)
        label_cell = ws[label_range.split(":")[0]]
        value_cell = ws[value_range.split(":")[0]]
        label_cell.value = label
        label_cell.fill = PatternFill("solid", fgColor=color)
        label_cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.value = formula
        value_cell.fill = PatternFill("solid", fgColor="F8FAFC")
        value_cell.font = Font(name="Aptos Display", size=22, bold=True, color=NAVY)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, bottom=THIN_GRAY)

    ws["A9"] = "Tiến độ review"
    ws["A9"].font = Font(name="Aptos Display", size=14, bold=True, color=NAVY)
    progress = [
        ["Trạng thái", "Số case"],
        ["PENDING", '=COUNTIF(\'Review\'!$B$2:$B$43,"PENDING")'],
        ["APPROVE", '=COUNTIF(\'Review\'!$B$2:$B$43,"APPROVE")'],
        ["CHANGE", '=COUNTIF(\'Review\'!$B$2:$B$43,"CHANGE")'],
        ["REJECT", '=COUNTIF(\'Review\'!$B$2:$B$43,"REJECT")'],
    ]
    for row in progress:
        ws.append(row)
    style_header(ws[10][0:2])
    ws["A11"].fill = PatternFill("solid", fgColor=LIGHT_AMBER)
    ws["A12"].fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    ws["A13"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["A14"].fill = PatternFill("solid", fgColor=LIGHT_RED)
    for row in range(11, 15):
        ws[f"B{row}"].number_format = "0"

    ws["D9"] = "Phủ Category cuối cùng"
    ws["D9"].font = Font(name="Aptos Display", size=14, bold=True, color=NAVY)
    ws["D10"] = "Code"
    ws["E10"] = "Tên Category"
    ws["F10"] = "Số case"
    style_header(ws[10][3:6])
    catalog = payload["catalog_snapshot"] + [{"code": "<null>", "display_name": "<null>"}]
    for row_index, item in enumerate(catalog, 11):
        ws.cell(row_index, 4, item["code"])
        ws.cell(row_index, 5, item["display_name"])
        ws.cell(row_index, 6, f'=COUNTIF(\'Review\'!$L$2:$L$43,E{row_index})')
    add_table(ws, f"D10:F{10 + len(catalog)}", "OverviewCoverageTable")

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Phân bố Category cuối cùng"
    chart.y_axis.title = "Category"
    chart.x_axis.title = "Số case"
    chart.height = 7.5
    chart.width = 14
    chart.add_data(Reference(ws, min_col=6, min_row=10, max_row=10 + len(catalog)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=5, min_row=11, max_row=10 + len(catalog)))
    chart.legend = None
    ws.add_chart(chart, "H9")

    ws["A17"] = "Cách review"
    ws["A17"].font = Font(name="Aptos Display", size=14, bold=True, color=NAVY)
    instructions = [
        "1. Mở sheet Review và lọc cột Cần xác nhận = CÓ trước.",
        "2. Chọn APPROVE nếu đồng ý toàn bộ nhãn; chọn CHANGE nếu cần sửa Category/Severity/Red flag.",
        "3. Khi chọn CHANGE, điền ba cột Reviewer và ghi lý do cụ thể.",
        "4. Không đổi trực tiếp expected_output trong sheet Chi tiết; sheet đó là ảnh chụp dữ liệu nguồn.",
        "5. Chỉ chuyển dataset thành golden chính thức sau khi mọi case được duyệt và bất đồng đã được phân xử.",
    ]
    for offset, text in enumerate(instructions, 18):
        ws.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=6)
        ws.cell(offset, 1, text)
        ws.cell(offset, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(offset, 1).font = Font(name="Aptos", size=10, color="344054")

    widths = {"A": 20, "B": 13, "C": 3, "D": 21, "E": 28, "F": 12, "G": 3, "H": 15, "I": 15, "J": 15, "K": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def build_review(wb: Workbook, cases: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Review")
    ws.sheet_view.showGridLines = False
    headers = [
        "Case ID",
        "Quyết định review",
        "Category reviewer",
        *(f"{label} (0-4)" for _key, label in CRITERIA),
        "Blocker reviewer",
        "Có ảnh",
        "Cần xác nhận",
        "Tiêu đề",
        "Ảnh",
        "Mô tả Cư dân",
        "Vị trí",
        "Category dự kiến",
        "Text category",
        "Image category",
        "Severity",
        "Red flag",
        "Image relevant",
        "Question kind",
        "Ghi chú cần duyệt",
        "File ảnh",
        "Nhận xét reviewer",
    ]
    ws.append(headers)
    style_header(ws[1])
    ws.row_dimensions[1].height = 38

    for row_index, item in enumerate(cases, 2):
        inputs = item["input"]
        output = item["expected_output"]
        image_paths = inputs["image_paths"]
        notes = "\n".join(item["review_notes"])
        row = [
            item["case_id"],
            "PENDING",
            None,
            None,
            None,
            "Có ảnh" if image_paths else "Text-only",
            "CÓ" if notes else "",
            item["title"],
            "",
            inputs["description"],
            " · ".join(filter(None, [inputs["location_label"], inputs["floor_label"], inputs.get("unit_code")])),
            output["category"],
            output["text_category"],
            output["image_category"],
            *(criterion_value(output, key) for key, _label in CRITERIA),
            ",".join(output.get("blockers") or []),
            output["image_relevant"],
            output["question_kind"],
            notes,
            "\n".join(image_paths),
            None,
        ]
        ws.append(row)
        ws.cell(row_index, 1).hyperlink = f"#'Chi tiết'!A{row_index}"
        ws.cell(row_index, 1).style = "Hyperlink"
        for col in range(1, len(headers) + 1):
            ws.cell(row_index, col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row_index, col).font = Font(name="Aptos", size=9, color="101828")
        if image_paths:
            source_path = REPO_ROOT / image_paths[0]
            image = fit_image(source_path)
            ws.add_image(image, f"I{row_index}")
            ws.cell(row_index, 20).hyperlink = str(source_path)
            ws.cell(row_index, 20).style = "Hyperlink"
            ws.row_dimensions[row_index].height = 68
        else:
            ws.row_dimensions[row_index].height = 45

    last_row = len(cases) + 1
    add_table(ws, f"A1:U{last_row}", "GoldenReviewTable")
    ws.freeze_panes = "H2"
    ws.auto_filter.ref = f"A1:U{last_row}"

    decision_dv = DataValidation(type="list", formula1='"PENDING,APPROVE,CHANGE,REJECT"', allow_blank=False)
    category_dv = DataValidation(type="list", formula1="'Catalog'!$B$2:$B$12", allow_blank=True)
    # 0-4, from a list rather than a free number: the scale has five steps and a
    # typed 7 is a judgement the rubric cannot read.
    criterion_dv = DataValidation(type="list", formula1='"0,1,2,3,4"', allow_blank=True)
    blocker_dv = DataValidation(type="list", formula1=f'"{",".join(BLOCKER_CODES)}"', allow_blank=True)
    for validation in (decision_dv, category_dv, criterion_dv, blocker_dv):
        ws.add_data_validation(validation)
    decision_dv.add(f"B2:B{last_row}")
    category_dv.add(f"C2:C{last_row}")
    # Columns D through H are the five criteria, I is the blocker.
    for column in "DEFGH":
        criterion_dv.add(f"{column}2:{column}{last_row}")
    blocker_dv.add(f"I2:I{last_row}")

    ws.conditional_formatting.add(
        f"B2:B{last_row}", FormulaRule(formula=["B2=\"APPROVE\""], fill=PatternFill("solid", fgColor=LIGHT_GREEN), font=Font(color=GREEN, bold=True))
    )
    ws.conditional_formatting.add(
        f"B2:B{last_row}", FormulaRule(formula=["B2=\"CHANGE\""], fill=PatternFill("solid", fgColor=LIGHT_BLUE), font=Font(color=BLUE, bold=True))
    )
    ws.conditional_formatting.add(
        f"B2:B{last_row}", FormulaRule(formula=["B2=\"REJECT\""], fill=PatternFill("solid", fgColor=LIGHT_RED), font=Font(color=RED, bold=True))
    )
    ws.conditional_formatting.add(
        f"G2:G{last_row}", FormulaRule(formula=["G2=\"CÓ\""], fill=PatternFill("solid", fgColor=LIGHT_AMBER), font=Font(color=AMBER, bold=True))
    )
    ws.conditional_formatting.add(
        f"P2:P{last_row}", FormulaRule(formula=["P2=TRUE"], fill=PatternFill("solid", fgColor=LIGHT_RED), font=Font(color=RED, bold=True))
    )
    ws.conditional_formatting.add(
        f"Q2:Q{last_row}", FormulaRule(formula=["Q2=FALSE"], fill=PatternFill("solid", fgColor=LIGHT_GRAY), font=Font(color=GRAY))
    )

    widths = [15, 18, 24, 18, 17, 11, 14, 34, 18, 48, 28, 23, 21, 21, 12, 11, 15, 25, 55, 48, 42]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = width
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def build_detail(wb: Workbook, cases: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Chi tiết")
    ws.sheet_view.showGridLines = False
    headers = [
        "case_id", "title", "review_status", "source_kind", "fixture_id", "input_mode", "image_state",
        "evidence_relation", "red_flag_source", "description", "image_paths", "location_label", "floor_label",
        "unit_code", "confirmed_category", "conversation", "category", "text_category", "image_category",
        *(key for key, _label in CRITERIA), "blockers",
        "understandable", "image_relevant", "location_consistent", "incident_facts",
        "ai_reason", "question_kind", "question_text", "category_options", "review_notes",
    ]
    ws.append(headers)
    style_header(ws[1])
    ws.row_dimensions[1].height = 36
    for item in cases:
        ws.append(flat_detail(item))
    last_row = len(cases) + 1
    add_table(ws, f"A1:AD{last_row}", "GoldenDetailTable")
    ws.freeze_panes = "J2"
    ws.auto_filter.ref = f"A1:AD{last_row}"
    for row in ws.iter_rows(min_row=2, max_row=last_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Aptos", size=9, color="101828")
        ws.row_dimensions[row[0].row].height = 48
    widths = {
        "A": 15, "B": 34, "C": 22, "D": 20, "E": 24, "F": 18, "G": 24, "H": 24, "I": 18,
        "J": 55, "K": 48, "L": 28, "M": 14, "N": 12, "O": 23, "P": 36, "Q": 24, "R": 21,
        "S": 21, "T": 12, "U": 11, "V": 15, "W": 16, "X": 19, "Y": 46, "Z": 60, "AA": 25,
        "AB": 55, "AC": 35, "AD": 60,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_catalog(wb: Workbook, payload: dict[str, Any], cases: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Catalog")
    ws.sheet_view.showGridLines = False
    style_title(ws, "A1:E1", "Category snapshot", payload["catalog_source"])
    ws.append([])
    ws.append(["Code", "Tên gửi cho model", "Số case Category cuối", "Base score", "Priority ceiling"])
    style_header(ws[4])
    counts = Counter(item["expected_output"]["category"] for item in cases)
    for item in payload["catalog_snapshot"]:
        ws.append([item["code"], item["display_name"], counts[item["display_name"]], "Xem migration", "Xem migration"])
    add_table(ws, f"A4:E{4 + len(payload['catalog_snapshot'])}", "CatalogSnapshotTable")
    for row in ws.iter_rows(min_row=5, max_row=4 + len(payload["catalog_snapshot"])):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    widths = {"A": 28, "B": 30, "C": 22, "D": 18, "E": 20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"


def build_confirmation(wb: Workbook, cases: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Cần xác nhận")
    ws.sheet_view.showGridLines = False
    style_title(ws, "A1:H1", "Các case cần xác nhận trước", "Được lấy từ trường review_notes của dataset nguồn")
    headers = ["Case ID", "Tiêu đề", "Ảnh", "Mô tả", "Category dự kiến", "An toàn con người", "Blocker", "Vấn đề cần xác nhận"]
    ws.append([])
    ws.append(headers)
    style_header(ws[4])
    filtered = [item for item in cases if item["review_notes"]]
    for row_index, item in enumerate(filtered, 5):
        inputs = item["input"]
        output = item["expected_output"]
        ws.append([
            item["case_id"], item["title"], "", inputs["description"], output["category"],
            criterion_value(output, "human_safety"), ",".join(output.get("blockers") or []),
            "\n".join(item["review_notes"]),
        ])
        ws.cell(row_index, 1).hyperlink = f"#'Review'!A{cases.index(item) + 2}"
        ws.cell(row_index, 1).style = "Hyperlink"
        if inputs["image_paths"]:
            image = fit_image(REPO_ROOT / inputs["image_paths"][0])
            ws.add_image(image, f"C{row_index}")
        for cell in ws[row_index]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Aptos", size=9, color="101828")
        ws.row_dimensions[row_index].height = 72
    add_table(ws, f"A4:H{4 + len(filtered)}", "ConfirmationTable")
    widths = {"A": 15, "B": 34, "C": 18, "D": 50, "E": 24, "F": 12, "G": 12, "H": 62}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "D5"


def build_guide(wb: Workbook) -> None:
    ws = wb.create_sheet("Hướng dẫn")
    ws.sheet_view.showGridLines = False
    style_title(ws, "A1:F1", "Hướng dẫn gán nhãn", "Áp dụng cho schema UnifiedClassification hiện tại")
    sections = [
        ("Category", "Chọn đúng một tên trong sheet Catalog. category là kết luận cuối; text_category và image_category chỉ là bằng chứng."),
        ("Severity", "LOW/MEDIUM/HIGH dựa trên phạm vi và ảnh hưởng thật. Không dùng LOW chỉ vì thiếu dữ liệu."),
        ("Red flag", "Chỉ bật khi có nguy hiểm trực tiếp ngay lúc này: khói, lửa, dây điện trần hở, ngập diện rộng, người bất tỉnh, người kẹt thang máy hoặc xô xát."),
        ("Image relevant", "TRUE khi ảnh cung cấp bằng chứng liên quan tới sự cố; FALSE khi gửi nhầm hoặc không đủ đọc. Đây là điểm cần chốt thêm cho ảnh mờ."),
        ("Question kind", "Chỉ dùng CATEGORY_CONFIRMATION, SEVERITY_CONFIRMATION hoặc LOCATION_CONFIRMATION khi thiếu đúng thông tin tương ứng; nếu không thì NONE."),
        ("APPROVE", "Dùng khi đồng ý Category, Severity, Red flag và hành vi hỏi lại."),
        ("CHANGE", "Điền Category/Severity/Red flag reviewer và lý do. Không sửa trực tiếp dữ liệu nguồn trong sheet Chi tiết."),
        ("REJECT", "Dùng khi case không thể gán nhãn tin cậy hoặc ảnh/input không phù hợp với mục tiêu eval."),
    ]
    ws["A4"] = "Mục"
    ws["B4"] = "Quy tắc review"
    style_header(ws[4][0:2])
    for row_index, (label, text) in enumerate(sections, 5):
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, text)
        ws.cell(row_index, 1).font = Font(name="Aptos", size=10, bold=True, color=NAVY)
        ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(row_index, 2).font = Font(name="Aptos", size=10, color="344054")
        for cell in ws[row_index]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_index].height = 42
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 95


def create_preview(path: Path, payload: dict[str, Any]) -> None:
    canvas = PILImage.new("RGB", (1500, 900), "white")
    draw = ImageDraw.Draw(canvas)
    title_font = ImageFont.truetype("arial.ttf", 34)
    header_font = ImageFont.truetype("arialbd.ttf", 20)
    small_font = ImageFont.truetype("arial.ttf", 14)
    draw.rectangle((0, 0, 1500, 95), fill="#17324D")
    draw.text((40, 24), "Golden Dataset V4 — Bảng review", font=title_font, fill="white")
    stats = [
        ("Tổng case", len(payload["cases"]), "#246BCE"),
        ("Case có ảnh", sum(bool(item["input"]["image_paths"]) for item in payload["cases"]), "#0F766E"),
        ("Red flag", sum(item["expected_output"]["red_flag"] for item in payload["cases"]), "#B42318"),
        ("Cần xác nhận", sum(bool(item["review_notes"]) for item in payload["cases"]), "#D97706"),
    ]
    for index, (label, value, color) in enumerate(stats):
        x = 40 + index * 355
        draw.rounded_rectangle((x, 125, x + 315, 245), radius=16, fill="#F8FAFC", outline=color, width=3)
        draw.text((x + 20, 143), label, font=header_font, fill=color)
        draw.text((x + 20, 182), str(value), font=title_font, fill="#17324D")
    draw.text((40, 285), "Preview sheet Review — các case cần xác nhận", font=header_font, fill="#17324D")
    columns = [(40, 150, "Case"), (190, 430, "Tiêu đề"), (620, 260, "Category"), (880, 120, "Severity"), (1000, 90, "RF"), (1090, 370, "Ghi chú")]
    y = 325
    for x, width, label in columns:
        draw.rectangle((x, y, x + width, y + 42), fill="#17324D")
        draw.text((x + 8, y + 10), label, font=small_font, fill="white")
    y += 42
    flagged = [item for item in payload["cases"] if item["review_notes"]][:7]
    for item in flagged:
        output = item["expected_output"]
        values = [
            item["case_id"],
            item["title"],
            output["category"] or "<null>",
            criterion_value(output, "human_safety") or "<null>",
            ",".join(output.get("blockers") or []) or "<none>",
            item["review_notes"][0],
        ]
        row_height = 68
        for (x, width, _), value in zip(columns, values):
            draw.rectangle((x, y, x + width, y + row_height), fill="#FFF4D6", outline="#D0D5DD", width=1)
            text = str(value)
            max_chars = max(8, int(width / 9))
            lines = [text[i : i + max_chars] for i in range(0, len(text), max_chars)][:3]
            for line_index, line in enumerate(lines):
                draw.text((x + 7, y + 8 + line_index * 17), line, font=small_font, fill="#101828")
        y += row_height
    canvas.save(path)


def build(output: Path, preview: Path) -> None:
    payload = json.loads(SOURCE.read_text(encoding="utf-8"))
    cases = payload["cases"]
    wb = Workbook()
    wb.properties.title = "Golden Dataset V4 Review"
    wb.properties.subject = payload["dataset_version"]
    wb.properties.creator = "Codex"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    build_overview(wb, payload)
    build_catalog(wb, payload, cases)
    build_review(wb, cases)
    build_detail(wb, cases)
    build_confirmation(wb, cases)
    build_guide(wb)

    wb._sheets = [
        wb["Tổng quan"], wb["Review"], wb["Cần xác nhận"], wb["Chi tiết"], wb["Catalog"], wb["Hướng dẫn"]
    ]
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    create_preview(preview, payload)

    check = load_workbook(output, data_only=False)
    assert check.sheetnames == ["Tổng quan", "Review", "Cần xác nhận", "Chi tiết", "Catalog", "Hướng dẫn"]
    assert check["Review"].max_row == len(cases) + 1
    assert check["Chi tiết"].max_row == len(cases) + 1
    assert len(check["Review"]._images) == 29
    assert len(check["Cần xác nhận"]._images) == 9
    assert check["Tổng quan"]["A5"].data_type == "f"
    check.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--preview", type=Path, default=DEFAULT_PREVIEW)
    args = parser.parse_args()
    build(args.output.resolve(), args.preview.resolve())
    print(args.output.resolve())


if __name__ == "__main__":
    main()
