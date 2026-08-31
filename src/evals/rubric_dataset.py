"""The 260-case rubric test set, read into the vocabulary the code uses.

`outputs/rubric-agent-eval-20260829/bo_test_agent_phan_loai_260_cases.xlsx` is
authored by hand, in Vietnamese, with its own column and value names. This
module is the one place that translates it. Everything downstream -- the
cross-check script, the tests, an eventual model run -- works in
`BlockerCode`, `Priority` and the five criterion names from
`src/domain/risk_scoring.py`.

**The workbook computes the rubric a second time.** `diem_rui_ro`,
`muc_theo_diem` and `muc_uu_tien_cuoi` are spreadsheet formulas over a `Rubric`
sheet that carries its own weights and thresholds. That is a reimplementation,
and reimplementations drift. `load_workbook_rubric` reads those constants back
so the drift is something a test can fail on rather than something a reviewer
has to notice across 260 rows.

Cached formula results are read with `data_only=True`, which returns whatever
Excel last stored. A workbook edited by a tool that does not recalculate will
have stale or absent cached values; `expected_risk_score` is `None` in that
case rather than zero, and the checker reports it as unevaluated rather than as
a mismatch.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from src.domain.risk_scoring import CRITERION_NAMES
from src.models.agent_schemas import BlockerCode
from src.models.enums import Priority

#: Where the workbook lives, relative to the repository root.
DEFAULT_DATASET = Path("outputs/rubric-agent-eval-20260829/bo_test_agent_phan_loai_260_cases.xlsx")

CASES_SHEET = "Test Cases"
RUBRIC_SHEET = "Rubric"
#: Two banner rows above the header row.
HEADER_ROW = 3

#: The workbook's blocker names, mapped onto the eleven contract codes.
#:
#: One-to-one and exhaustive in both directions, which `tests/test_evals/`
#: asserts. A dataset code with no contract code would be a case measuring a
#: floor the system cannot apply.
BLOCKER_CODE_BY_LOCAL_NAME: dict[str, BlockerCode] = {
    "CHAY_HOAC_KHOI_DANG_XAY_RA": BlockerCode.FIRE_OR_SMOKE,
    "DIEN_HO_DANG_MANG_DIEN": BlockerCode.ELECTRIC_SHOCK_OR_LIVE_WIRE,
    "RO_GAS_HOAC_NGAT_KHI": BlockerCode.GAS_LEAK_OR_ASPHYXIATION,
    "CHAN_THUONG_NGHIEM_TRONG": BlockerCode.SERIOUS_INJURY,
    "NGUOI_KET_TRONG_THANG_MAY": BlockerCode.PERSON_TRAPPED_IN_ELEVATOR,
    "LOI_THOAT_DUY_NHAT_BI_CHAN": BlockerCode.SOLE_ESCAPE_ROUTE_BLOCKED,
    "BAO_LUC_TRUC_TIEP_DANG_XAY_RA": BlockerCode.ONGOING_VIOLENCE,
    "NUOC_THAI_DANG_TRAO": BlockerCode.SEWAGE_OVERFLOW,
    "NUOC_CHAY_MANH_LAN_SANG_CAN_KHAC": BlockerCode.HEAVY_WATER_FLOW_SPREAD_RISK,
    "MAT_HOAN_TOAN_DIEN_KHONG_THEO_KE_HOACH": BlockerCode.TOTAL_UNPLANNED_UTILITY_LOSS,
    "TOILET_DUY_NHAT_KHONG_SU_DUNG_DUOC": BlockerCode.SOLE_TOILET_UNUSABLE,
}

#: `tieu_chi_muc_tieu` -- which criterion a clarifying question is aimed at.
CRITERION_BY_LOCAL_NAME: dict[str, str] = {
    "NGUY_CO_AN_TOAN": "human_safety",
    "THIET_HAI_VA_LAN_RONG": "property_spread",
    "MAT_CHUC_NANG_THIET_YEU": "essential_function",
    "PHAM_VI": "affected_scope",
    "TOC_DO_XAU_DI": "deterioration_speed",
}

#: The five score columns, in contract order.
CRITERION_COLUMNS: dict[str, str] = {
    "human_safety": "diem_nguy_co_an_toan",
    "property_spread": "diem_thiet_hai_va_lan_rong",
    "essential_function": "diem_mat_chuc_nang_thiet_yeu",
    "affected_scope": "diem_pham_vi_agent",
    "deterioration_speed": "diem_toc_do_xau_di",
}

#: `trang_thai_phan_loai`: what the Agent is expected to conclude.
READY = "SAN_SANG_KET_LUAN"
NEEDS_CLARIFICATION = "CAN_LAM_RO"
REJECTED = "TU_CHOI"

#: `trang_thai_review`: whether a human has signed the row off.
REVIEW_APPROVED = "DA_DUYET"


@dataclass(frozen=True)
class RubricCase:
    """One row, in the code's vocabulary.

    Fields the Agent is expected to produce and fields the backend is expected
    to derive are kept apart, because they fail for different reasons: the first
    is a model-quality result and the second is an arithmetic one.
    """

    tc_id: str
    group: str
    description: str
    #: What the Agent should conclude: READY / NEEDS_CLARIFICATION / REJECTED.
    classification_state: str
    #: The five 0-4 scores, or None where the row does not expect a score.
    criteria: dict[str, int | None]
    blockers: tuple[BlockerCode, ...]
    #: Which criterion a clarifying question should target, if any.
    question_criteria: tuple[str, ...]
    #: Backend side.
    confirmed_unit_count: int | None
    scope_source: str | None
    effective_scope_score: int | None
    expected_risk_score: Decimal | None
    expected_score_priority: Priority | None
    expected_final_priority: Priority | None
    emergency_gate: str | None
    assignment_allowed: bool | None
    grouping_allowed: bool | None
    review_state: str | None

    @property
    def criteria_complete(self) -> bool:
        return all(self.criteria.get(name) is not None for name in CRITERION_NAMES)

    @property
    def human_reviewed(self) -> bool:
        return self.review_state == REVIEW_APPROVED


@dataclass(frozen=True)
class WorkbookRubric:
    """The rubric constants the workbook computes with, for drift checks."""

    weights: dict[str, Decimal]
    #: Lower bound of each band, ascending.
    thresholds: dict[Priority, Decimal]
    blocker_floors: dict[BlockerCode, Priority]


def _rows(path: Path, sheet: str) -> list[tuple]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook[sheet].iter_rows(values_only=True))
    finally:
        workbook.close()


def _as_int(value) -> int | None:
    if value is None or isinstance(value, str) and not value.strip():
        return None
    if isinstance(value, bool):
        return None
    return int(value)


def _as_decimal(value) -> Decimal | None:
    """Cached formula results only. A formula string means Excel never stored
    a value for it, which is not the same as a value of zero."""
    if value is None or isinstance(value, str):
        return None
    return Decimal(str(value)).quantize(Decimal("0.01"))


def _as_priority(value) -> Priority | None:
    if not isinstance(value, str) or not value.strip():
        return None
    try:
        return Priority(value.strip())
    except ValueError:
        return None


def _split(value) -> tuple[str, ...]:
    if not isinstance(value, str) or not value.strip():
        return ()
    return tuple(item.strip() for item in value.split(";") if item.strip())


def load_rubric_cases(path: Path | str = DEFAULT_DATASET) -> list[RubricCase]:
    """Every case in the workbook, translated.

    An unrecognised blocker name raises rather than being dropped: a case
    measuring a floor the system has no code for is a case nobody can act on,
    and silently skipping it would make the dataset look smaller than it is.
    """
    rows = _rows(Path(path), CASES_SHEET)
    header = rows[HEADER_ROW - 1]
    index = {name: position for position, name in enumerate(header) if isinstance(name, str)}

    def field(row, name):
        position = index.get(name)
        return None if position is None else row[position]

    cases: list[RubricCase] = []
    for row in rows[HEADER_ROW:]:
        tc_id = field(row, "tc_id")
        if not isinstance(tc_id, str) or not tc_id.strip():
            continue
        local_blockers = _split(field(row, "ma_dieu_kien_chan"))
        unknown = [name for name in local_blockers if name not in BLOCKER_CODE_BY_LOCAL_NAME]
        if unknown:
            raise ValueError(f"{tc_id}: blocker name(s) not in the contract: {unknown}")
        question_criteria = tuple(
            CRITERION_BY_LOCAL_NAME[name]
            for name in _split(field(row, "tieu_chi_muc_tieu"))
            if name in CRITERION_BY_LOCAL_NAME
        )
        cases.append(
            RubricCase(
                tc_id=tc_id.strip(),
                group=str(field(row, "cum_kiem_thu") or ""),
                description=str(field(row, "mo_ta_nguoi_dung") or ""),
                classification_state=str(field(row, "trang_thai_phan_loai") or ""),
                criteria={
                    name: _as_int(field(row, column)) for name, column in CRITERION_COLUMNS.items()
                },
                blockers=tuple(BLOCKER_CODE_BY_LOCAL_NAME[name] for name in local_blockers),
                question_criteria=question_criteria,
                confirmed_unit_count=_as_int(field(row, "so_can_da_xac_nhan")),
                scope_source=field(row, "nguon_diem_pham_vi"),
                effective_scope_score=_as_int(field(row, "diem_pham_vi_ap_dung")),
                expected_risk_score=_as_decimal(field(row, "diem_rui_ro")),
                expected_score_priority=_as_priority(field(row, "muc_theo_diem")),
                expected_final_priority=_as_priority(field(row, "muc_uu_tien_cuoi")),
                emergency_gate=field(row, "trang_thai_p5_gate"),
                assignment_allowed=field(row, "cho_phep_phan_viec"),
                grouping_allowed=field(row, "cho_phep_grouping"),
                review_state=field(row, "trang_thai_review"),
            )
        )
    return cases


def load_workbook_rubric(path: Path | str = DEFAULT_DATASET) -> WorkbookRubric:
    """The weights, thresholds and floors the spreadsheet formulas use."""
    rows = _rows(Path(path), RUBRIC_SHEET)
    weight_labels = {
        "Nguy cơ an toàn con người": "human_safety",
        "Thiệt hại tài sản / lan rộng": "property_spread",
        "Mất chức năng sinh hoạt thiết yếu": "essential_function",
        "Phạm vi unit bị ảnh hưởng": "affected_scope",
        "Tốc độ xấu đi": "deterioration_speed",
    }
    weights: dict[str, Decimal] = {}
    thresholds: dict[Priority, Decimal] = {}
    floors: dict[BlockerCode, Priority] = {}
    for row in rows:
        label = row[0]
        if not isinstance(label, str):
            continue
        label = label.strip()
        if label in weight_labels and row[1] is not None:
            weights[weight_labels[label]] = Decimal(str(row[1]))
        elif label in {band.value for band in Priority} and row[1] is not None:
            thresholds[Priority(label)] = Decimal(str(row[1]))
        elif label in BLOCKER_CODE_BY_LOCAL_NAME:
            band = _as_priority(row[1])
            if band is not None:
                floors[BLOCKER_CODE_BY_LOCAL_NAME[label]] = band
    return WorkbookRubric(weights=weights, thresholds=thresholds, blocker_floors=floors)


__all__ = [
    "BLOCKER_CODE_BY_LOCAL_NAME",
    "CRITERION_BY_LOCAL_NAME",
    "CRITERION_COLUMNS",
    "DEFAULT_DATASET",
    "NEEDS_CLARIFICATION",
    "READY",
    "REJECTED",
    "REVIEW_APPROVED",
    "RubricCase",
    "WorkbookRubric",
    "load_rubric_cases",
    "load_workbook_rubric",
]
